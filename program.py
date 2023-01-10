import csv
import requests


import numpy as np
from matplotlib import pyplot as plt
from jinja2 import Environment, FileSystemLoader

import doctest
import unittest

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side

from os import listdir
from os.path import isfile, join
from multiprocessing import Process, Queue, Manager

from unittest import TestCase

import pdfkit

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

URl = 'https://www.cbr.ru/scripts/XML_valFull.asp?d=0&date_req='

def csv_parser(file_name):
    """Функция читает информацию из файла и создает объекты vacancy,
        которые складывает в два словаря:
            по годам,
            по городам

    :param file_name: Имя csv файла с данными
    :return:
    />>> csv_parser('vacancies_by_year.csv')[0]['2007'][0].salary_currency
    'RUR'
    """

    with open(file_name, encoding='utf_8_sig') as r_file:
        file_reader = csv.reader(r_file, delimiter=",")
        list_naming = file_reader.__next__()
        vacancies, vacancies_city = {}, {}
        for row in file_reader:
            # if len(row) != len(list_naming) or row.__contains__(""):
            #     continue

            if len(row) == 0:
                continue
            fields = {}
            for field_index in range(len(row)):
                fields.update({list_naming[field_index]: row[field_index]})
            year = fields["published_at"][:4]
            vacancy = Vacancy(name=fields["name"],
                              salary_from=fields["salary_from"],
                              salary_to=fields["salary_to"],
                              salary_currency=fields["salary_currency"],
                              area_name=fields["area_name"],
                              published_at=fields["published_at"])
            if vacancy.salary_from == 0 and vacancy.salary_to == 0:
                continue
            if vacancies_city.keys().__contains__(vacancy.area_name):
                vacancies_city[vacancy.area_name].append(vacancy)
            else:
                vacancies_city.update({vacancy.area_name: [vacancy]})
            if vacancies.keys().__contains__(year):
                vacancies[year].append(vacancy)
            else:
                vacancies.update({year: [vacancy]})

        return vacancies, vacancies_city
def throw_low_quantity_currencies(vacancies):
    if vacancies is None or len(vacancies) == 0:
        return
    currencies_rate = {}
    for key in vacancies.keys():
        for vacancy in vacancies[key]:
            if vacancy.salary_currency in currencies_rate:
                currencies_rate[vacancy.salary_currency] += 1
            else:
                currencies_rate.update({vacancy.salary_currency: 1})
    """сохраняем валюты кооторые встречаются более 5000 раз"""
    currencies_to_pop = []
    for currency in currencies_rate:
        if currencies_rate[currency] <= 5000:
            currencies_to_pop.append(currency)
    for currency in currencies_to_pop:
        currencies_rate.pop(currency)


    """Убираем вакансии валюты которых встречались реже тербуемого"""
    min_date, max_date = '2100-12-30', '0001-01-01'
    for key in vacancies.keys():
        vacancies_to_pop = []
        for vacancy in vacancies[key]:
            if vacancy.salary_currency not in currencies_rate.keys():
                vacancies_to_pop.append(vacancy)
            else:
                min_date = min(min_date, vacancy.published_at)
                max_date = max(max_date, vacancy.published_at)
        for vacancy in vacancies_to_pop:
            vacancies[key].remove(vacancy)


    from test import generate_currencies_dataframe

    date = min_date.split('-')
    min_date = '/'.join(date[::-1])
    date = max_date.split('-')
    max_date = '/'.join(date[::-1])
    if min_date == '2100-12-30' and max_date == '0001-01-01':
        return
    dataframe = generate_currencies_dataframe(list(currencies_rate.keys()), min_date, max_date)

    for key in vacancies.keys():
        for vacancy in vacancies[key]:
            vacancy.transfer_currency(dataframe)

    return vacancies

"""
TODO : написать функцию котторая будет считать количестов вакансий для существующих валют
написать функцию которая бы запрашивала курс для указанной валюты в указанную дату
написать функцию которая переводит зарплату из собственной валюты в рубли"""

class VacancyTests(TestCase):
    """Этот класс тестирует коректность инициализации класса Vacancy

    """
    def test_vacancy_type(self):
        self.assertEqual(type(Vacancy('Программист', 10, 20, 'USD', 'Moscow', '2007-12-03T17:34:36+0300')).__name__,
                         'Vacancy')

    def test_salary_average(self):
        self.assertEqual(Vacancy('Программист', 10, 20, 'USD', 'Moscow', '2007-12-03T17:34:36+0300').salary_average,
                         15.0)

    def test_salary_currency(self):
        self.assertEqual(Vacancy('Программист', 10, 20, 'USD', 'Moscow', '2007-12-03T17:34:36+0300').salary_currency,
                         'USD')
class CsvParserTests(TestCase):
    """Этот класс тестирует работу метода csv_parser на заранее определенно верных примерах
    """
    def test_salary_currency(self):
        self.assertEqual(csv_parser('vacancies_by_year.csv')[0]['2007'][0].salary_currency, 'RUR')

    def test_salary_average(self):
        self.assertEqual(csv_parser('vacancies_by_year.csv')[0]['2007'][0].salary_average, 40000.0)


class Vacancy:
    """Класс для представления вакансии.
    Attributes:
        name            (str):  Наименование вакансии
        salary_from     (int):  Нижняя граница оклада
        salary_to       (int):  Верхняя граница оклада
        salary_average  (int):  Среднее значение оклада
        salary_currency (str):  Валюта оклада
        area_name       (str):  Название региона вакансии
        published_at    (str):  Дата публикации вакансии
    """

    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at):
        """Инициализирует объект Vacancy, выполняет рассчёт среднего значения оклада

        :param name            (str):  Наименование вакансии
        :param salary_from     (int):  Нижняя граница оклада
        :param salary_to       (int):  Верхняя граница оклада
        :param salary_currency (str):  Валюта оклада
        :param area_name       (str):  Название региона вакансии
        :param published_at    (str):  Дата публикации вакансии

        >>> type(Vacancy('Программист', 10, 20, 'USD', 'Moscow', '2007-12-03T17:34:36+0300')).__name__
        'Vacancy'
        >>> Vacancy('Программист', 10, 20, 'USD', 'Moscow', '2007-12-03T17:34:36+0300').salary_average
        15.0
        >>> Vacancy('Программист', 10, 20, 'USD', 'Moscow', '2007-12-03T17:34:36+0300').salary_currency
        'USD'
        """

        self.name = name
        self.salary_from = float(salary_from) if salary_from != '' else 0
        self.salary_to = float(salary_to) if salary_to != '' else 0
        self.salary_average = (self.salary_from + self.salary_to) / 2
        self.salary_currency = salary_currency
        self.area_name = area_name
        self.published_at = published_at

    def get_date_for_request(self):
        """функия которая возвращают дату публикации вакансии для url запроса
        >>> Vacancy('Программист', 10, 20, 'USD', 'Moscow', '2007-12-03T17:34:36+0300').get_date_for_request()
        '03/12/2007'
        """

        date = self.published_at[:10].split('-')
        return '01' + ('/'.join(date[::-1]))[2:]

    def transfer_currency(self, data_frame):
        if data_frame[self.get_date_for_request()].__contains__(self.salary_currency):
            k = data_frame[self.get_date_for_request()][self.salary_currency]
        else:
            k = 0
        self.salary_from = self.salary_from * k
        self.salary_to = self.salary_to * k
        self.salary_average = self.salary_average * k


class Report:
    """Класс для формирования отчёта

        Attributes:
            font_size (int): размер шрифта в отчёте
    """

    font_size = 12

    def generate_png(self, statics_by_years, statics_by_cities):
        """Функция генерирующая графики по статистике
        Args:
            :param statics_by_years: словарь, содержащий значения типа:
                {год: [средняя зарплата за год, средняя зарплата за год для указанной вакансии,
                 количество вакансий за год, количество указанных вакансий за год]}

            :param statics_by_cities: словарь, содержащий значения типа:
            {город: [средняя зарплата по городу, процент вакансий в городе]}
        """

        fig, axes = plt.subplots(nrows=2, ncols=2, figsize=(18, 12))
        plt.xticks(fontsize=8)
        years = np.arange(len(statics_by_years.keys()))
        cities = np.arange(len(statics_by_cities))

        colors = ['#FFAC1C', '#FF0000', '#800000', '#FFFF00', '#808000', '#00FF00', '#00FFFF', '#008080', '#0000FF',
                  '#FF00FF', '#800080']

        axes[0][0].bar(years, [x[0] for x in statics_by_years.values()], width=0.4, label="средняя з/п")
        axes[0][0].bar([x + 0.4 for x in years], [x[1] for x in statics_by_years.values()], width=0.4,
                       label="средняя з/п")
        axes[0][0].set_xticks(years, statics_by_years.keys(), rotation=90, ha='right')
        axes[0][0].yaxis.grid(True)
        axes[0][0].set_title("Уровень зарплаты по годам")
        axes[0][0].legend(loc='upper left')

        axes[0][1].bar(years, [x[2] for x in statics_by_years.values()], width=0.4, label="Количество вакансий")
        axes[0][1].bar([x + 0.4 for x in years], [x[3] for x in statics_by_years.values()], width=0.4,
                       label=f"Количество вакансий{name}")
        axes[0][1].set_xticks(years, statics_by_years.keys(), rotation=90, ha='right')
        axes[0][1].yaxis.grid(True)
        axes[0][1].set_title("Количество вакансий по годам")
        axes[0][1].legend(loc='upper left')

        statics_by_cities = dict(sorted(statics_by_cities.items(), key=lambda item: item[1][0], reverse=True))

        axes[1][0].barh(cities, [x[0] for x in statics_by_cities.values()])
        axes[1][0].set_yticks(cities)
        axes[1][0].set_yticklabels(statics_by_cities.keys())
        axes[1][0].invert_yaxis()  # labels read top-to-bottom
        axes[1][0].xaxis.grid(True)
        axes[1][0].set_title('Уровень зарплаты по годам')

        statics_by_cities = dict(sorted(statics_by_cities.items(), key=lambda item: item[1][1], reverse=True))

        axes[1][1].pie(list([x[1] for x in (list(statics_by_cities.values()))[:10]]) + list(
            [sum([x[1] for x in (list(statics_by_cities.values()))[10:]])]),
                       labels=(list(statics_by_cities.keys()))[:10] + ['Другие'], colors=colors)

        fig.savefig('graph.png')
        plt.close(fig)

    def generate_excel(self, statics_by_years, statics_by_cities):
        """Функция генерирующая эксель таблицу по статистике

        Args:
            :param statics_by_years: словарь, содержащий значения типа:
                {год: [средняя зарплата за год, средняя зарплата за год для указанной вакансии,
                 количество вакансий за год, количество указанных вакансий за год]}

            :param statics_by_cities: словарь, содержащий значения типа:
                {город: [средняя зарплата по городу, процент вакансий в городе]}

        """

        wb = Workbook()
        wb.remove(wb['Sheet'])

        statistics_by_year_sheet = wb.create_sheet("Статистика по годам")
        statistics_by_year_sheet.append(["Год", "Средняя зарплата", f"Средняя зарплата - {name}", "Количество вакансий",
                                         f"Количество вакансий - {name}"])
        for key in statics_by_years:
            statistics_by_year_sheet.append([key] + statics_by_years[key])

        cols_dict = {}
        for row in statistics_by_year_sheet.rows:
            for cell in row:
                letter = cell.column_letter
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                if cell.value:
                    cell.font = Font(name='Calibri', size=self.font_size)
                    len_cell = len(str(cell.value))
                    len_cell_dict = 0
                    if letter in cols_dict:
                        len_cell_dict = cols_dict[letter]

                    if len_cell > len_cell_dict:
                        cols_dict[letter] = len_cell
                        ###!!! ПРОБЛЕМА АВТОМАТИЧЕСКОЙ ПОДГОНКИ !!!###
                        ###!!! расчет новой ширины колонки (здесь надо подгонять) !!!###
                        new_width_col = len_cell * self.font_size ** (self.font_size * 0.009)
                        statistics_by_year_sheet.column_dimensions[cell.column_letter].width = new_width_col

        for cell in statistics_by_year_sheet[1:1]:
            cell.font = Font(name='Calibri', size=self.font_size, bold=True)

        statistics_by_cities_sheet = wb.create_sheet("Статистика по городам")
        statistics_by_cities_sheet.append(['Город', 'Уровень зарплат', ' ', 'Город', 'Доля вакансий'])
        for key in statics_by_cities:
            statistics_by_cities_sheet.append([key, statics_by_cities[key][0], ' ', key, statics_by_cities[key][1]])

        cols_dict = {}
        for row in statistics_by_cities_sheet.rows:
            for cell in row:
                letter = cell.column_letter
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                if cell.value:
                    cell.font = Font(name='Calibri', size=self.font_size)
                    len_cell = len(str(cell.value))
                    len_cell_dict = 0
                    if letter in cols_dict:
                        len_cell_dict = cols_dict[letter]

                    if len_cell > len_cell_dict:
                        cols_dict[letter] = len_cell
                        ###!!! ПРОБЛЕМА АВТОМАТИЧЕСКОЙ ПОДГОНКИ !!!###
                        ###!!! расчет новой ширины колонки (здесь надо подгонять) !!!###
                        new_width_col = len_cell * self.font_size ** (self.font_size * 0.009)
                        statistics_by_cities_sheet.column_dimensions[cell.column_letter].width = new_width_col

        for cell in statistics_by_cities_sheet['E']:
            cell.number_format = '0.00%'
        for cell in statistics_by_cities_sheet[1:1]:
            cell.font = Font(name='Calibri', size=self.font_size, bold=True)
        wb.save('report.xlsx')

    def generate_report(self, statics_by_years, statics_by_cities):
        """Функция генерирующая пдф файл с отчётом по статистике

        Args:
            :param statics_by_years: словарь, содержащий значения типа:
                {год: [средняя зарплата за год, средняя зарплата за год для указанной вакансии,
                 количество вакансий за год, количество указанных вакансий за год]}

            :param statics_by_cities: словарь, содержащий значения типа:
                {город: [средняя зарплата по городу, процент вакансий в городе]}

        """
        self.generate_png(statics_by_years, statics_by_cities)
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        config = pdfkit.configuration(wkhtmltopdf=r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe")

        pdf_template = template.render({'items': statics_by_years, 'items2': statics_by_cities})
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'out.pdf', options=options, configuration=config)


def union_dict(dict1, dict2):
    """Метод дополняет пары ключ-значение второго словаря в первый

    :param dict1: Словарь, который будем дополнять
    :param dict2: Добавочный словарь
    :return: dict1 + dict2
    """
    if dict2 is None or dict1 is None:
        return
    for key2 in dict2:
        if dict1.keys().__contains__(key2):
            dict1[key2] += dict2[key2]
        else:
            dict1.update({key2: dict2[key2]})
    return dict1

def get_statistic_by_years(vacancies_by_years, name):
    """Этот метод подсчитывает среднюю зарплату для всех вакансий и указанных, а так же их количество

    :param vacancies_by_years: список вакансий
    :param name: имя вакансии
    :return:
    """
    statistics_by_years = {}
    vacancies_count_by_years, vacancies_salary_by_years, \
    vacancies_count_by_years_for_name, vacancies_salary_by_years_for_name = {}, {}, {}, {}
    for key in vacancies_by_years.keys():
        vacancies_count_by_years.update({key: len(vacancies_by_years[key])})
        vacancies_salary_by_years.update({key: int(
            sum(x.salary_average for x in vacancies_by_years[key]) / vacancies_count_by_years[
                key])})
        vacancies_count_by_years_for_name.update({key: len(list(filter(lambda x: (name in x.name), vacancies_by_years[key])))})
        if vacancies_count_by_years_for_name[key] == 0:
            vacancies_salary_by_years_for_name.update({key: 0})
        else:
            vacancies_salary_by_years_for_name.update({key: int(sum(
                x.salary_average for x in
                list(filter(lambda x: (name in x.name), vacancies_by_years[key]))) / vacancies_count_by_years_for_name[key])})
        statistics_by_years.update({key: [vacancies_salary_by_years[key], vacancies_salary_by_years_for_name[key],
                                          vacancies_count_by_years[key], vacancies_count_by_years_for_name[key]]})
    return statistics_by_years

def get_statistic_by_cities(vacancies_city):
    """Этот метод отбрасывает данные о городах, вакансии которых занимают меньше одного процента от рынка,
    а так же вычисляет средние зарплаты и процент на рынке для остальных городов.

    :param vacancies_city: словарь типа {city: [vacancy1, vacancy2, ...]}
    :return: статистика по городам
    """
    vacancies_count = sum(len(vacancies_city[i]) for i in vacancies_city)
    city_to_pop = []
    for city in vacancies_city:
        if (len(vacancies_city[city]) < int(vacancies_count / 100)):
            city_to_pop.append(city)

    for city in city_to_pop:
        vacancies_city.pop(city)

    map(lambda k, v: (k, v) if len(vacancies_city[k]) > vacancies_count / 100 else None, vacancies_city)

    statistics_by_cities = {}
    vacancies_salary_by_city, vacancies_proportion_by_city = {}, {}
    for key in vacancies_city.keys():
        vacancies_salary_by_city.update(
            {key: int(
                sum(map(lambda x: x.salary_average, vacancies_city[key])) / len(
                    vacancies_city[key]))})
        vacancies_proportion_by_city.update({key: float(len(vacancies_city[key]) / vacancies_count)})
        statistics_by_cities.update(
            {key: [vacancies_salary_by_city[key], round(vacancies_proportion_by_city[key] * 100, 2)]})

    return statistics_by_cities

def generate_statistic_by_years(vacancies_by_years, name, q):
    """Метод который обрабатывается отдельным процессом для каждого файла,
    вычисляет статистику по годам, и сохраняет данные по городам.

    :param filename:
    :param name:
    :param q:
    :return:
    """

    q.put(get_statistic_by_years(vacancies_by_years, name))

def read_chunks(filename, q):
    parsed = csv_parser(filename)
    q.put(parsed)

if __name__ == '__main__':
    filename, name = input("Введите название файла: "), input("Введите название профессии: ")

    """При пустом вводе имени файла подставляет стандартное имя файла
    """
    if filename == "":
        filename = "data/vacancies"

    """Запускает новый поток для обработки каждого файла в папке
    """
    files = [f for f in listdir(filename) if isfile(join(filename, f))]
    processes = []
    statistic_by_city = {}
    for file in files:
        manager = Manager()
        q = manager.Queue()
        # if os is win then change / to \\
        p = Process(target=read_chunks, args=(filename+'/'+file, q))
        processes.append((p, q))
        p.start()

    # region
    """Обработка данных по городам и формирование статистики по ним
    """
    statistic_by_years = {}
    vacancies_by_years = {}
    vacancies_by_cities = {}
    for i in range(len(processes)):
        processes[i][0].join()
        data = processes[i][1].get()
        vacancies_by_years = union_dict(vacancies_by_years, data[0])
        vacancies_by_cities = union_dict(vacancies_by_cities, data[1])

    vacancies_by_years = throw_low_quantity_currencies(vacancies_by_years)

    processes.clear()
    for year in vacancies_by_years.keys():
        manager = Manager()
        q = manager.Queue()
        # if os is win then change / to \\
        p = Process(target=generate_statistic_by_years, args=(vacancies_by_years, name,q))
        processes.append((p, q))
        p.start()

    for process in processes:
        process[0].join()
        data = process[1].get()

        # data_for_year = processes[i][1].get()
        # data_for_cities = processes[i][1].get()
        #
        # statistic_by_city = union_dict(statistic_by_city, data_for_cities)
        statistic_by_years.update(data)

    statistics_by_cities = get_statistic_by_cities(vacancies_by_cities)
    vacancies_salary_by_city = dict(map(lambda x: (x, statistics_by_cities[x][0]), statistics_by_cities))
    vacancies_proportion_by_city = dict(map(lambda x: (x, statistics_by_cities[x][1]), statistics_by_cities))
    # endregion
    # region
    """Блок кода для вывода данных в консоль
    """
    temp = '\''
    print(f"Динамика уровня зарплат по годам: "
          + str(dict(map(lambda x: (x, statistic_by_years[x][0]),
                         statistic_by_years))).replace(temp, ''))
    print("Динамика количества вакансий по годам: "
          + str(dict(map(lambda x: (x, statistic_by_years[x][2]),
                         statistic_by_years))).replace(temp, ''))
    print("Динамика уровня зарплат по годам для выбранной профессии: "
          + str(dict(map(lambda x: (x, statistic_by_years[x][1]),
                         statistic_by_years))).replace(temp, ''))
    print("Динамика количества вакансий по годам для выбранной профессии: "
          + str(dict(map(lambda x: (x, statistic_by_years[x][3]),
                         statistic_by_years))).replace(temp, ''))

    vacancies_salary_by_city = {k: v for k, v in
                                sorted(vacancies_salary_by_city.items(),
                                       key=lambda item: item[1],
                                       reverse=True)
                                [0:10]}
    vacancies_proportion_by_city = {k: round(v, 4) for k, v in
                                    sorted(vacancies_proportion_by_city.items(),
                                           key=lambda item: item[1],
                                           reverse=True)[
                                    0:10]}

    print("Уровень зарплат по городам (в порядке убывания): " + str(vacancies_salary_by_city))
    print(f"Доля вакансий по городам (в порядке убывания): {str(vacancies_proportion_by_city)}")
    # endregion
